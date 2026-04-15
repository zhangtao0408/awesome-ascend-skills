import argparse
import os
import re
from collections import defaultdict, namedtuple
from io import BytesIO

import matplotlib.pyplot as plt
import pandas as pd
from decimal import Decimal, getcontext
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# excel表格最大列宽
MAX_COL_WIDTH = 20
# task_summary_by_process_pid表每张图中最多展示20个pid，其余部分进行拆分
MAX_PID_PER_CHART = 20
# task_summary_by_process_pid表每种comm图的起始列
BASE_COL_INDEX = 10
ROW_HEIGHT = 15

# 定义判断阈值
CS_COUNT_THRESHOLD = 10000  # 上下文切换次数阈值
MIGRATE_COUNT_THRESHOLD = 100  # CPU迁移次数阈值
IRQ_TIME_THRESHOLD = 0.1  # 中断时间阈值（秒）

TaskKey = namedtuple("TaskKey", ["cpu", "comm", "pid"])

CPU_RE = re.compile(r"\[(\d+)]")
TS_RE = re.compile(r"\s(\d+\.\d+):")
IRQ_ENTRY_RE = re.compile(r"irq_handler_entry: irq=(\d+)\s+name=(\S+)")
IRQ_EXIT_RE = re.compile(r"irq_handler_exit: irq=(\d+)\s+ret=")
SOFT_ENTRY_RE = re.compile(r"softirq_entry: vec=(\d+)\s+\[action=([^]]+)]")
SOFT_EXIT_RE = re.compile(r"softirq_exit: vec=(\d+)\s+\[action=([^]]+)]")
WAKING_RE = re.compile(
    r"sched_waking: comm=([^ ]+) pid=(\d+) prio=\d+ target_cpu=(\d+)"
)
WAKEUP_RE = re.compile(
    r"sched_wakeup: comm=([^ ]+) pid=(\d+) prio=\d+ success=\d+ target_cpu=(\d+)"
)
SWITCH_RE = re.compile(
    r"sched_switch: prev_comm=([^ ]+) prev_pid=(\d+) prev_prio=\d+ prev_state=([A-Z0-9]+) "
    r"==> next_comm=([^ ]+) next_pid=(\d+) next_prio=\d+"
)
MIGRATE_RE = re.compile(
    r"sched_migrate_task: comm=([^ ]+) pid=(\d+) prio=\d+ orig_cpu=(\d+) dest_cpu=(\d+)"
)


class TaskStats:
    __slots__ = ("running", "sleeping", "runnable", "cs_count", "migrate_count", "irqs")

    def __init__(self):
        self.running = Decimal("0")
        self.sleeping = Decimal("0")
        self.runnable = Decimal("0")
        self.cs_count = 0
        self.migrate_count = 0
        self.irqs = defaultdict(lambda: {"count": 0, "time": Decimal("0")})


class OffCPUState:
    __slots__ = ("state", "timestamp", "cpu")

    def __init__(self, state, timestamp, cpu):
        self.state = state
        self.timestamp = timestamp
        self.cpu = cpu


class IRQFrame:
    __slots__ = ("irq_type", "irq_id", "name", "start_ts", "task_key", "acc_time")

    def __init__(self, irq_type, irq_id, name, start_ts, task_key):
        self.irq_type = irq_type
        self.irq_id = irq_id
        self.name = name
        self.start_ts = start_ts
        self.task_key = task_key
        self.acc_time = Decimal("0")


class TraceAnalyzer:
    def __init__(self):
        self.current_task = {}
        self.last_ts = defaultdict(lambda: None)
        self.irq_stack = defaultdict(list)
        self.stats = defaultdict(TaskStats)
        self.offcpu = {}
        self.pid_comm = {}

    @staticmethod
    def classify_prev_state(prev_state_str: str) -> str:
        if any(ch in prev_state_str for ch in ("S", "D", "K", "W", "T")):
            return "sleeping"
        elif "R" in prev_state_str:
            return "runnable"
        else:
            return "idle"

    @staticmethod
    def get_task_key(cpu: int, comm: str, pid: int) -> TaskKey:
        return TaskKey(cpu=cpu, comm=comm, pid=int(pid))

    def advance_time(self, cpu: int, timestamp: Decimal) -> None:
        last_timestamp = self.last_ts[cpu]
        if not last_timestamp or timestamp < last_timestamp:
            self.last_ts[cpu] = timestamp
            return
        dt = timestamp - last_timestamp
        if self.irq_stack[cpu]:
            self.irq_stack[cpu][-1].acc_time += dt
        elif cpu in self.current_task:
            comm, pid = self.current_task[cpu]
            if pid != 0:
                self.stats[self.get_task_key(cpu, comm, pid)].running += dt
        self.last_ts[cpu] = timestamp

    def handle_soft_entry(
        self, cpu: int, timestamp: Decimal, vec: str, action: str
    ) -> None:
        self.advance_time(cpu, timestamp)
        comm, pid = self.current_task.get(cpu, (f"swapper/{cpu}", 0))
        self.irq_stack[cpu].append(
            IRQFrame(
                "soft", int(vec), action, timestamp, self.get_task_key(cpu, comm, pid)
            )
        )

    def handle_soft_exit(self, cpu: int, timestamp: Decimal) -> None:
        self.advance_time(cpu, timestamp)
        if self.irq_stack[cpu]:
            frame = self.irq_stack[cpu].pop()
            if frame.task_key.pid != 0:
                s = self.stats[frame.task_key].irqs[("soft", frame.irq_id, frame.name)]
                s["count"] += 1
                s["time"] += frame.acc_time

    def handle_wakeup(self, timestamp: Decimal, comm: str, pid: str, cpu: str) -> None:
        pid = int(pid)
        cpu = int(cpu)
        self.pid_comm[pid] = comm
        if pid in self.offcpu and self.offcpu[pid].state == "sleeping":
            return
        self.offcpu[pid] = OffCPUState("runnable", timestamp, cpu)

    def handle_sched_switch(self, cpu: int, timestamp: Decimal, m) -> None:
        prev_comm, prev_pid, prev_state, next_comm, next_pid = m.groups()
        prev_pid, next_pid = int(prev_pid), int(next_pid)
        self.advance_time(cpu, timestamp)
        self.pid_comm[prev_pid] = prev_comm
        self.pid_comm[next_pid] = next_comm
        if prev_pid != 0:
            self.offcpu[prev_pid] = OffCPUState(
                self.classify_prev_state(prev_state), timestamp, cpu
            )
            self.stats[self.get_task_key(cpu, prev_comm, prev_pid)].cs_count += 1
        if next_pid != 0 and next_pid in self.offcpu:
            oc = self.offcpu.pop(next_pid)
            dt = timestamp - oc.timestamp
            if dt > 0:
                comm = self.pid_comm.get(next_pid, next_comm)
                key = self.get_task_key(cpu, comm, next_pid)
                if oc.state == "sleeping":
                    self.stats[key].sleeping += dt
                elif oc.state == "runnable":
                    self.stats[key].runnable += dt
        self.current_task[cpu] = (next_comm, next_pid)

    def handle_migrate(self, m) -> None:
        comm, pid, orig_cpu, _ = m.groups()
        self.stats[self.get_task_key(int(orig_cpu), comm, int(pid))].migrate_count += 1

    def handle_irq_entry(
        self, cpu: int, timestamp: Decimal, irq_id: str, name: str
    ) -> None:
        self.advance_time(cpu, timestamp)
        comm, pid = self.current_task.get(cpu, (f"swapper/{cpu}", 0))
        self.irq_stack[cpu].append(
            IRQFrame(
                "hard", int(irq_id), name, timestamp, self.get_task_key(cpu, comm, pid)
            )
        )

    def handle_irq_exit(self, cpu: int, timestamp: Decimal) -> None:
        self.advance_time(cpu, timestamp)
        if self.irq_stack[cpu]:
            frame = self.irq_stack[cpu].pop()
            if frame.task_key.pid != 0:
                s = self.stats[frame.task_key].irqs[("hard", frame.irq_id, frame.name)]
                s["count"] += 1
                s["time"] += frame.acc_time

    def parse_trace_file(self, path: str) -> Decimal:
        trace_end_ts = Decimal("0")
        getcontext().prec = 30
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                m_cpu = CPU_RE.search(line)
                m_ts = TS_RE.search(line)
                if not m_cpu or not m_ts:
                    continue
                cpu = int(m_cpu.group(1))
                timestamp = Decimal(m_ts.group(1))
                trace_end_ts = timestamp
                if "softirq_entry" in line:
                    m = SOFT_ENTRY_RE.search(line)
                    if m:
                        self.handle_soft_entry(cpu, timestamp, *m.groups())
                    continue
                if "softirq_exit" in line:
                    m = SOFT_EXIT_RE.search(line)
                    if m:
                        self.handle_soft_exit(cpu, timestamp)
                    continue
                if "sched_wakeup" in line:
                    m = WAKEUP_RE.search(line)
                    if m:
                        self.handle_wakeup(timestamp, *m.groups())
                    continue
                if "sched_waking" in line:
                    m = WAKING_RE.search(line)
                    if m:
                        self.handle_wakeup(timestamp, *m.groups())
                    continue
                if "sched_switch" in line:
                    m = SWITCH_RE.search(line)
                    if m:
                        self.handle_sched_switch(cpu, timestamp, m)
                    continue
                if "sched_migrate_task" in line:
                    m = MIGRATE_RE.search(line)
                    if m:
                        self.handle_migrate(m)
                    continue
                if "irq_handler_entry" in line:
                    m = IRQ_ENTRY_RE.search(line)
                    if m:
                        self.handle_irq_entry(cpu, timestamp, *m.groups())
                    continue
                if "irq_handler_exit" in line:
                    m = IRQ_EXIT_RE.search(line)
                    if m:
                        self.handle_irq_exit(cpu, timestamp)
                    continue
                self.advance_time(cpu, timestamp)
        return trace_end_ts

    @staticmethod
    def write_excel(
        df_summary,
        df_irq_detail,
        df_proc_irq,
        df_summary_by_comm,
        df_summary_by_process_pid,
        path: str,
    ) -> None:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="task_summary", index=False)
            df_summary_by_comm.to_excel(
                writer, sheet_name="task_summary_by_comm", index=False
            )
            df_summary_by_process_pid.to_excel(
                writer, sheet_name="task_summary_by_process_pid", index=False
            )
            df_irq_detail.to_excel(writer, sheet_name="task_irq_detail", index=False)
            df_proc_irq.to_excel(writer, sheet_name="proc_irq_detail", index=False)

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    cell_value = str(cell.value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                ws.column_dimensions[col_letter].width = min(
                    max_length + 2, MAX_COL_WIDTH
                )
        wb.save(path)

    @staticmethod
    def insert_charts(
        excel_path: str,
        df_summary_by_comm: pd.DataFrame,
        df_summary_by_process_pid: pd.DataFrame,
        df_proc_irq: pd.DataFrame,
    ) -> None:
        wb = load_workbook(excel_path)

        # ---------- 图 1：Top10 running_s -> task_summary_by_comm!I2 ----------
        ws_comm = wb["task_summary_by_comm"]
        df_top10_comm = df_summary_by_comm.sort_values(
            "running_s", ascending=False
        ).head(10)
        plt.figure(figsize=(6, 4))
        ax1 = plt.gca()
        ax1.bar(df_top10_comm["comm"], df_top10_comm["running_s"], color="steelblue")
        ax1.set_ylabel("Running Time (s)")
        ax1.set_title("Top 10 Processes by Running Time")
        ax1.tick_params(axis="x", rotation=45)
        ax1.margins(y=0.1)
        for x, y in zip(df_top10_comm["comm"], df_top10_comm["running_s"]):
            ax1.text(x, y, f"{int(y)}", ha="center", va="bottom", fontsize=8)
        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=120)
        plt.close()
        buf.seek(0)
        img = Image(buf)
        ws_comm.add_image(img, "I2")

        # ---------- 图 2：每个 comm 的 PID 指标 -> task_summary_by_process_pid ----------
        ws_pid = wb["task_summary_by_process_pid"]
        insert_row = 2
        chart_col_index = BASE_COL_INDEX
        df_summary_by_process_pid = df_summary_by_process_pid.copy()
        df_summary_by_process_pid["irq_total_time_ms"] = (
            df_summary_by_process_pid["irq_total_time_s"] * 1000
        )

        metrics = [
            ("cs_count", "Context Switch Count"),
            ("migrate_count", "CPU Migration Count"),
            ("irq_total_time_ms", "IRQ Total Time (ms)"),
        ]

        for comm in df_top10_comm["comm"]:
            df_proc = df_summary_by_process_pid[
                df_summary_by_process_pid["comm"] == comm
            ].copy()
            df_proc = df_proc.sort_values("pid")
            rows = df_proc.to_dict("records")
            chunks = [
                rows[i : i + MAX_PID_PER_CHART]
                for i in range(0, len(rows), MAX_PID_PER_CHART)
            ]

            for chunk_index, chunk in enumerate(chunks):
                fig, axes = plt.subplots(3, 1, figsize=(8, 6))
                fig.suptitle(
                    f"{comm} - PID Metrics ({chunk_index + 1}/{len(chunks)})",
                    fontsize=12,
                )

                for ax, (metric, title) in zip(axes, metrics):
                    pids = [str(r["pid"]) for r in chunk]
                    values = [r[metric] for r in chunk]

                    ax.bar(pids, values, color="royalblue")
                    ax.set_title(title, fontsize=10)
                    ax.set_ylabel(metric)
                    ax.tick_params(axis="x", labelsize=7, rotation=45)
                    ax.margins(y=0.15)

                    for x, y in zip(pids, values):
                        ax.text(x, y, f"{int(y)}", ha="center", va="bottom", fontsize=6)

                plt.tight_layout(rect=[0, 0, 1, 0.95])
                buf = BytesIO()
                plt.savefig(buf, format="png", dpi=100)
                plt.close()
                buf.seek(0)

                img = Image(buf)
                col_letter = get_column_letter(chart_col_index)
                ws_pid.add_image(img, f"{col_letter}{insert_row}")

                chart_col_index += 12  # 横向向右移动

            # 每个 comm 结束后换一行，列重置
            insert_row += int(img.height / ROW_HEIGHT) + 2
            chart_col_index = BASE_COL_INDEX

        # ---------- 图 3：IRQ Top10 -> proc_irq_detail!H2 ----------
        ws_proc_irq = wb["proc_irq_detail"]
        df_proc_irq = df_proc_irq.copy()
        df_proc_irq["comm_pid"] = (
            df_proc_irq["comm"] + "_" + df_proc_irq["pid"].astype(str)
        )
        df_proc_irq["irq_time_ms"] = df_proc_irq["irq_time_s"] * 1000
        df_top10_irq_time = df_proc_irq.sort_values(
            "irq_time_ms", ascending=False
        ).head(10)
        df_top10_irq_count = df_proc_irq.sort_values("irq_count", ascending=False).head(
            10
        )
        fig, axes = plt.subplots(2, 1, figsize=(10, 6))
        ax_t = axes[0]
        ax_t.bar(
            df_top10_irq_time["comm_pid"],
            df_top10_irq_time["irq_time_ms"],
            color="purple",
        )
        ax_t.set_title("Top 10 IRQ Time by Process", fontsize=10)
        ax_t.set_ylabel("IRQ Time (ms)")
        ax_t.tick_params(axis="x", labelsize=8, rotation=45)
        ax_t.margins(y=0.1)
        for x, y in zip(
            df_top10_irq_time["comm_pid"], df_top10_irq_time["irq_time_ms"]
        ):
            ax_t.text(x, y, f"{int(y)}", ha="center", va="bottom", fontsize=6)
        ax_c = axes[1]
        ax_c.bar(
            df_top10_irq_count["comm_pid"],
            df_top10_irq_count["irq_count"],
            color="teal",
        )
        ax_c.set_title("Top 10 IRQ Count by Process", fontsize=10)
        ax_c.set_ylabel("IRQ Count")
        ax_c.tick_params(axis="x", labelsize=8, rotation=45)
        ax_c.margins(y=0.1)
        for x, y in zip(
            df_top10_irq_count["comm_pid"], df_top10_irq_count["irq_count"]
        ):
            ax_c.text(x, y, f"{int(y)}", ha="center", va="bottom", fontsize=6)
        plt.tight_layout()
        buf = BytesIO()
        plt.savefig(buf, format="png", dpi=100)
        plt.close()
        buf.seek(0)
        img = Image(buf)
        ws_proc_irq.add_image(img, "H2")

        wb.save(excel_path)

    def build_dataframes(self):
        summary_rows = []
        irq_detail_rows = []
        proc_irq = defaultdict(lambda: {"count": 0, "time": Decimal("0")})
        for tk, s in self.stats.items():
            irq_total_count = sum(v["count"] for v in s.irqs.values())
            irq_total_time = sum(v["time"] for v in s.irqs.values())
            summary_rows.append(
                {
                    "cpu": tk.cpu,
                    "comm": tk.comm,
                    "pid": tk.pid,
                    "running_s": s.running,
                    "sleeping_s": s.sleeping,
                    "runnable_s": s.runnable,
                    "cs_count": s.cs_count,
                    "migrate_count": s.migrate_count,
                    "irq_total_count": irq_total_count,
                    "irq_total_time_s": irq_total_time,
                }
            )
            for (irq_type, irq_id, name), v in s.irqs.items():
                irq_detail_rows.append(
                    {
                        "cpu": tk.cpu,
                        "comm": tk.comm,
                        "pid": tk.pid,
                        "irq_type": irq_type,
                        "irq_id": irq_id,
                        "irq_name": name,
                        "irq_count": v["count"],
                        "irq_time_s": v["time"],
                    }
                )
                key = (tk.comm, tk.pid, irq_type, irq_id, name)
                proc_irq[key]["count"] += v["count"]
                proc_irq[key]["time"] += v["time"]

        df_summary = pd.DataFrame(summary_rows)
        df_irq_detail = pd.DataFrame(irq_detail_rows)
        df_proc_irq = pd.DataFrame(
            [
                {
                    "comm": comm,
                    "pid": pid,
                    "irq_type": irq_type,
                    "irq_id": irq_id,
                    "irq_name": name,
                    "irq_count": v["count"],
                    "irq_time_s": v["time"],
                }
                for (comm, pid, irq_type, irq_id, name), v in proc_irq.items()
            ]
        )
        summary_by_comm = defaultdict(
            lambda: {
                "running_s": Decimal("0"),
                "sleeping_s": Decimal("0"),
                "runnable_s": Decimal("0"),
                "cs_count": 0,
                "migrate_count": 0,
                "irq_total_count": 0,
                "irq_total_time_s": Decimal("0"),
            }
        )
        for row in summary_rows:
            key = row["comm"]
            summary_by_comm[key]["running_s"] += row["running_s"]
            summary_by_comm[key]["sleeping_s"] += row["sleeping_s"]
            summary_by_comm[key]["runnable_s"] += row["runnable_s"]
            summary_by_comm[key]["cs_count"] += row["cs_count"]
            summary_by_comm[key]["migrate_count"] += row["migrate_count"]
            summary_by_comm[key]["irq_total_count"] += row["irq_total_count"]
            summary_by_comm[key]["irq_total_time_s"] += row["irq_total_time_s"]
        df_summary_by_comm = pd.DataFrame(
            [{"comm": comm, **v} for comm, v in summary_by_comm.items()]
        )
        df_summary_by_process_pid = df_summary.groupby(
            ["comm", "pid"], as_index=False
        ).agg(
            {
                "running_s": "sum",
                "sleeping_s": "sum",
                "runnable_s": "sum",
                "cs_count": "sum",
                "migrate_count": "sum",
                "irq_total_count": "sum",
                "irq_total_time_s": "sum",
            }
        )
        for df in (
            df_summary,
            df_irq_detail,
            df_proc_irq,
            df_summary_by_comm,
            df_summary_by_process_pid,
        ):
            for col in df.columns:
                if df[col].dtype == "object":
                    df[col] = df[col].apply(
                        lambda x: float(x) if isinstance(x, Decimal) else x
                    )
        return (
            df_summary,
            df_irq_detail,
            df_proc_irq,
            df_summary_by_comm,
            df_summary_by_process_pid,
        )


def analyze_top10_processes_by_comm(df_summary_by_comm, df_irq_detail):
    """按进程名分析top10进程的性能指标是否存在问题"""
    # 按照运行时间排序，获取top10进程
    df_top10 = df_summary_by_comm.sort_values("running_s", ascending=False).head(10)

    print("\n=== Top 10 进程性能分析 (按进程名) ===")

    # 检查是否存在问题的标志
    has_host_issue = False

    # 检查每个进程的指标
    for _, row in df_top10.iterrows():
        print(f"\n进程: {row['comm']}")
        print(f"  运行时间: {row['running_s']:.2f}秒")
        print(f"  上下文切换次数: {row['cs_count']}")
        print(f"  CPU迁移次数: {row['migrate_count']}")
        print(f"  中断总时间: {row['irq_total_time_s']:.2f}秒")

        # 判断是否存在问题
        issues = []
        if row["cs_count"] > CS_COUNT_THRESHOLD:
            issues.append(f"上下文切换次数过多 (> {CS_COUNT_THRESHOLD})")
        if row["migrate_count"] > MIGRATE_COUNT_THRESHOLD:
            issues.append(f"CPU迁移次数过多 (> {MIGRATE_COUNT_THRESHOLD})")
        if row["irq_total_time_s"] > IRQ_TIME_THRESHOLD:
            issues.append(f"中断时间过长 (> {IRQ_TIME_THRESHOLD}秒)")

        if issues:
            print(f"  问题: {'; '.join(issues)}")
            has_host_issue = True

            # 如果中断时间过长，找出打断该进程时间最长的几个中断
            if row["irq_total_time_s"] > IRQ_TIME_THRESHOLD:
                print(f"\n  打断该进程时间最长的5个中断:")
                # 筛选该进程的中断记录
                process_irqs = df_irq_detail[df_irq_detail["comm"] == row["comm"]]
                # 按中断时间排序，取前5个
                top_irqs = process_irqs.sort_values("irq_time_s", ascending=False).head(
                    5
                )
                for _, irq_row in top_irqs.iterrows():
                    print(
                        f"    - {irq_row['irq_type']} IRQ {irq_row['irq_id']} ({irq_row['irq_name']}): {irq_row['irq_time_s']:.2f}秒 ({irq_row['irq_count']}次)"
                    )
        else:
            print(f"  状态: 正常")

    return has_host_issue


def run(args: argparse.Namespace) -> None:
    if not os.path.exists(args.input):
        print(f"输入文件 {args.input} 不存在")
        return

    tracer = TraceAnalyzer()
    trace_end_ts = tracer.parse_trace_file(args.input)

    try:
        # 处理剩余时间
        for cpu in tracer.last_ts.keys():
            tracer.advance_time(cpu, trace_end_ts)

        # 构建数据框
        (
            df_summary,
            df_irq_detail,
            df_proc_irq,
            df_summary_by_comm,
            df_summary_by_process_pid,
        ) = tracer.build_dataframes()

        # 分析top10进程（按进程名）
        has_issue = analyze_top10_processes_by_comm(df_summary_by_comm, df_irq_detail)

        # 分析所有存在问题的进程
        print("\n=== 所有存在问题的进程分析 ===")
        problematic_processes = df_summary[
            (df_summary["cs_count"] > CS_COUNT_THRESHOLD)
            | (df_summary["migrate_count"] > MIGRATE_COUNT_THRESHOLD)
            | (df_summary["irq_total_time_s"] > IRQ_TIME_THRESHOLD)
        ]

        if not problematic_processes.empty:
            print(f"检测到 {len(problematic_processes)} 个存在问题的进程:")
            for _, row in problematic_processes.iterrows():
                print(f"- 进程: {row['comm']} (PID: {row['pid']})")
                issues = []
                if row["cs_count"] > CS_COUNT_THRESHOLD:
                    issues.append(
                        f"上下文切换次数过高: {row['cs_count']} (阈值: {CS_COUNT_THRESHOLD})"
                    )
                if row["migrate_count"] > MIGRATE_COUNT_THRESHOLD:
                    issues.append(
                        f"CPU迁移次数过高: {row['migrate_count']} (阈值: {MIGRATE_COUNT_THRESHOLD})"
                    )
                if row["irq_total_time_s"] > IRQ_TIME_THRESHOLD:
                    issues.append(
                        f"中断时间过长: {row['irq_total_time_s']:.2f}s (阈值: {IRQ_TIME_THRESHOLD}s)"
                    )
                print("  " + "; ".join(issues))

            # 输出最终结论
            print("\n=== 分析结论 ===")
            print(
                "❌ 发现Host侧性能问题: 存在进程上下文切换、CPU迁移或中断时间异常的情况"
            )
            print("   建议进一步分析相关进程，优化Host侧性能")
        else:
            # 输出最终结论
            print("\n=== 分析结论 ===")
            print("✅ 未发现明显的Host侧性能问题")

        # 生成Excel文件
        output_path = args.output
        if not output_path:
            # 如果未指定输出路径，在当前路径下生成
            input_filename = os.path.basename(args.input)
            output_path = f"{os.path.splitext(input_filename)[0]}_analysis.xlsx"

        # 保存基本统计表格
        tracer.write_excel(
            df_summary,
            df_irq_detail,
            df_proc_irq,
            df_summary_by_comm,
            df_summary_by_process_pid,
            output_path,
        )
        print(f"\n✅ 基本统计表格已生成: {output_path}")

        # 插入图表
        tracer.insert_charts(
            output_path, df_summary_by_comm, df_summary_by_process_pid, df_proc_irq
        )
        print("✅ 图表插入完成")
        print(f"\n🎉 分析结果已保存到Excel文件: {output_path}")

    except Exception as e:
        print(f"分析过程中发生错误: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Trace Analyzer - Host侧性能分析")
    parser.add_argument("--input", type=str, required=True, help="输入trace文件路径")
    parser.add_argument(
        "--output", type=str, help="输出Excel文件路径（可选，默认在当前路径生成）"
    )
    args = parser.parse_args()
    run(args)
